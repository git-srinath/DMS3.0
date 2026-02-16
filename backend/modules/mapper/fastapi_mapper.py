from typing import Any, Dict, List, Optional
import datetime
import io
import traceback
import re
import difflib

import pandas as pd
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.database.dbconnect import (
    create_metadata_connection,
    create_target_connection,
)
from backend.modules.helper_functions import (
    get_mapping_ref,
    get_mapping_details,
    get_parameter_mapping_datatype,
    get_parameter_mapping_scd_type,
    check_if_job_already_created,
    create_update_mapping,
    create_update_mapping_detail,
    validate_all_mapping_details,
    get_error_messages_list,
    call_activate_deactivate_mapping,
    call_delete_mapping,
    call_delete_mapping_details,
)
# Import validate_logic2 directly from pkgdwmapr_python to support target_connection parameter
from backend.modules.mapper.pkgdwmapr_python import validate_logic2

# Support both FastAPI (package import) and legacy Flask (relative import) contexts
try:
    from backend.modules.logger import info, error, warning
except ImportError:  # When running Flask app.py directly inside backend
    from modules.logger import info, error, warning  # type: ignore


router = APIRouter(tags=["mapper"])

# Constants for form and table fields (matching Flask mapper.py)
FORM_FIELDS = [
    "reference",
    "description",
    "sourceSystem",
    "tableName",
    "tableType",
    "targetSchema",
    "freqCode",
    "bulkProcessRows",
]
TABLE_FIELDS = [
    "primaryKey",
    "pkSeq",
    "fieldName",
    "dataType",
    "fieldDesc",
    "scdType",
    "keyColumn",
    "valColumn",
    "logic",
    "mapCombineCode",
    "execSequence",
]


def _detect_db_type_from_connection(conn) -> str:
    """
    Lightweight database type detection based on connection module.
    Duplicated here to avoid circular imports.
    """
    module_name = type(conn).__module__
    if "psycopg" in module_name or "pg8000" in module_name:
        return "POSTGRESQL"
    if "oracledb" in module_name or "cx_Oracle" in module_name:
        return "ORACLE"
    # Default to ORACLE for backward compatibility
    return "ORACLE"


def _normalize_sql(sql: str) -> str:
    """
    Normalize SQL for comparison:
    - Remove comments
    - Collapse whitespace
    - Uppercase
    """
    if not sql:
        return ""

    # Remove single-line comments --
    sql_no_single = re.sub(r"--.*?$", "", sql, flags=re.MULTILINE)
    # Remove /* */ comments
    sql_no_comments = re.sub(r"/\*.*?\*/", "", sql_no_single, flags=re.DOTALL)
    # Collapse whitespace
    sql_spaced = re.sub(r"\s+", " ", sql_no_comments).strip()
    return sql_spaced.upper()


def _calculate_sql_similarity(a: str, b: str) -> float:
    """
    Calculate similarity between two SQL strings using difflib.
    Returns a float between 0 and 1.
    """
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def _build_wrapped_sql(original_sql: str, db_type: str) -> str:
    """
    Wrap the user SQL so we can safely fetch column metadata without
    loading a large result set.
    """
    original_sql = original_sql.strip().rstrip(";")
    if not original_sql:
        raise ValueError("SQL content is empty")

    # Use WHERE 1=0 to avoid fetching data while still getting metadata
    if db_type == "POSTGRESQL":
        return f"SELECT * FROM ({original_sql}) AS subq WHERE 1=0"
    # Oracle and others
    return f"SELECT * FROM ({original_sql}) subq WHERE 1=0"


@router.get("/get-parameter-mapping-datatype")
async def get_parameter_mapping_datatype_api() -> List[Dict[str, Any]]:
    """
    Return parameter mapping entries of type 'Datatype'.
    Mirrors Flask endpoint: GET /mapper/get-parameter-mapping-datatype
    """
    conn = None
    try:
        conn = create_metadata_connection()
        return get_parameter_mapping_datatype(conn)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error in get_parameter_mapping_datatype: {str(e)}",
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@router.get("/parameter_scd_type")
async def parameter_scd_type() -> List[Dict[str, Any]]:
    """
    Return parameter mapping entries of type 'SCD'.
    Mirrors Flask endpoint: GET /mapper/parameter_scd_type
    """
    conn = None
    try:
        conn = create_metadata_connection()
        try:
            parameter_data = get_parameter_mapping_scd_type(conn)
            return parameter_data
        finally:
            conn.close()
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error in parameter_scd_type: {str(e)}"
        )


@router.get("/get-connections")
async def get_connections() -> List[Dict[str, Any]]:
    """
    Get list of active database connections from DMS_DBCONDTLS.
    Mirrors Flask endpoint: GET /mapper/get-connections
    """
    conn = None
    try:
        conn = create_metadata_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT conid, connm, dbhost, dbsrvnm, usrnm
                FROM DMS_DBCONDTLS
                WHERE curflg = 'Y'
                ORDER BY connm
            """
            )

            connections: List[Dict[str, Any]] = []
            for row in cursor.fetchall():
                connections.append(
                    {
                        "conid": str(row[0]),
                        "connm": row[1],
                        "dbhost": row[2],
                        "dbsrvnm": row[3],
                        "usrnm": row[4] if len(row) > 4 else None,
                    }
                )

            cursor.close()
            return connections
        finally:
            conn.close()
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error fetching connections: {str(e)}"
        )


@router.get("/get-by-reference/{reference}")
async def get_by_reference(reference: str):
    """
    Fetch mapping header + detail rows by reference.
    Mirrors Flask endpoint: GET /mapper/get-by-reference/<reference>
    """
    conn = None
    try:
        conn = create_metadata_connection()
        try:
            # Get reference data
            main_result: Optional[Dict[str, Any]] = get_mapping_ref(conn, reference)

            if not main_result:
                raise HTTPException(
                    status_code=404,
                    detail={
                        "exists": False,
                        "message": (
                            "Reference not found or inactive. You can create a new "
                            "mapping with this reference."
                        ),
                    },
                )

            # Get mapping details
            details_result: List[Dict[str, Any]] = get_mapping_details(conn, reference)

            # Get job created status
            job_status = check_if_job_already_created(conn, reference)

            # Format response
            form_data = {
                "reference": main_result.get("MAPREF"),
                "description": main_result.get("MAPDESC") or "",
                "mapperId": str(main_result.get("MAPID")),
                "targetSchema": main_result.get("TRGSCHM") or "",
                "tableName": main_result.get("TRGTBNM") or "",
                "tableType": main_result.get("TRGTBTYP") or "",
                "freqCode": main_result.get("FRQCD") or "",
                "sourceSystem": main_result.get("SRCSYSTM") or "",
                "bulkProcessRows": main_result.get("BLKPRCROWS"),
                "targetConnectionId": (
                    str(main_result.get("TRGCONID"))
                    if main_result.get("TRGCONID")
                    else None
                ),
                "isReferenceDisabled": True,
                "logic_verification_status": main_result.get("LGVRFYFLG"),
                "activate_status": main_result.get("STFLG"),
                "job_creation_status": job_status,
                # Checkpoint configuration (handle NULL values for older mappings)
                "checkpointStrategy": (
                    main_result.get("CHKPNTSTRTGY")
                    if main_result.get("CHKPNTSTRTGY")
                    else "AUTO"
                ),
                "checkpointColumn": (
                    main_result.get("CHKPNTCLNM")
                    if main_result.get("CHKPNTCLNM")
                    else ""
                ),
                "checkpointEnabled": (
                    main_result.get("CHKPNTENBLD") == "Y"
                    if main_result.get("CHKPNTENBLD")
                    else True
                ),
                # Optional: persisted base/source SQL text for this mapping (if BASESQL/SRCSQL column exists)
                "baseSql": (
                    main_result.get("BASESQL")
                    or main_result.get("basesql")
                    or main_result.get("SRCSQL")
                    or ""
                ),
            }

            # Transform the details result into rows
            rows: List[Dict[str, Any]] = []
            for row in details_result:
                rows.append(
                    {
                        "mapdtlid": str(row.get("MAPDTLID")),
                        "mapref": row.get("MAPREF") or "",
                        "fieldName": row.get("TRGCLNM") or "",
                        "dataType": row.get("TRGCLDTYP") or "",
                        "primaryKey": row.get("TRGKEYFLG") == "Y",
                        "pkSeq": (
                            str(row.get("TRGKEYSEQ"))
                            if row.get("TRGKEYSEQ") is not None
                            else ""
                        ),
                        "fieldDesc": row.get("TRGCLDESC") or "",
                        "logic": row.get("MAPLOGIC") or "",
                        "keyColumn": row.get("KEYCLNM") or "",
                        "valColumn": row.get("VALCLNM") or "",
                        "mapCombineCode": row.get("MAPCMBCD") or "",
                        "execSequence": (
                            str(row.get("EXCSEQ"))
                            if row.get("EXCSEQ") is not None
                            else ""
                        ),
                        "scdType": (
                            str(row.get("SCDTYP"))
                            if row.get("SCDTYP") is not None
                            else ""
                        ),
                        "LogicVerFlag": row.get("LGVRFYFLG"),
                    }
                )

            # If no detail rows exist, provide empty template rows
            if not rows:
                rows = [
                    {
                        "mapdtlid": "",
                        "mapref": reference,
                        "fieldName": "",
                        "dataType": "",
                        "primaryKey": False,
                        "pkSeq": "",
                        "fieldDesc": "",
                        "logic": "",
                        "keyColumn": "",
                        "valColumn": "",
                        "mapCombineCode": "",
                        "execSequence": "",
                        "scdType": "",
                        "LogicVerFlag": "",
                    }
                    for _ in range(6)
                ]

            response_data = {
                "exists": True,
                "formData": form_data,
                "rows": rows,
                "message": "Mapping data retrieved successfully",
            }

            return response_data
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail={
                "error": "An error occurred while retrieving the mapping data",
                "details": str(e),
            },
        )


# ----- Write / validation / activation endpoints -----


class SaveToDbRequest(BaseModel):
    formData: Dict[str, Any]
    rows: List[Dict[str, Any]]
    modifiedRows: Optional[List[int]] = None


@router.post("/save-to-db")
async def save_to_db(payload: SaveToDbRequest):
    """
    Save mapping header and detail rows.
    Mirrors Flask endpoint: POST /mapper/save-to-db
    """
    conn = None
    try:
        data = payload.model_dump()
        form_data = data["formData"]
        rows = data["rows"]
        modified_rows = data.get("modifiedRows") or []
        user_id = form_data.get("username")

        conn = create_metadata_connection()
        try:
            target_connection_id = form_data.get("targetConnectionId")

            checkpoint_strategy = form_data.get("checkpointStrategy", "AUTO")
            checkpoint_column = form_data.get("checkpointColumn", None)
            checkpoint_enabled = (
                "Y" if form_data.get("checkpointEnabled", True) else "N"
            )

            # Get targetSchema with fallback - auto-populate from connection if missing
            target_schema = form_data.get("targetSchema", "").strip() if form_data.get("targetSchema") else ""
            
            # If targetSchema is empty but targetConnectionId is provided, fetch username from connection
            if not target_schema and target_connection_id:
                try:
                    cursor = conn.cursor()
                    db_type = _detect_db_type_from_connection(conn)
                    # Use table name directly - DMS_DBCONDTLS should work for both Oracle and PostgreSQL
                    # For PostgreSQL, we'll let the database handle case sensitivity
                    table_name = "DMS_DBCONDTLS"
                    
                    if db_type == "POSTGRESQL":
                        query = f'SELECT usrnm FROM "{table_name}" WHERE conid = %s AND curflg = %s'
                        cursor.execute(query, (target_connection_id, 'Y'))
                    else:  # Oracle
                        query = f"SELECT usrnm FROM {table_name} WHERE conid = :conid AND curflg = 'Y'"
                        cursor.execute(query, {"conid": target_connection_id})
                    
                    result = cursor.fetchone()
                    cursor.close()
                    
                    if result and result[0]:
                        target_schema = str(result[0]).strip().upper()
                except Exception as e:
                    warning(f"Could not fetch username from connection {target_connection_id}: {str(e)}")
                    # Continue with empty target_schema - validation will handle it
            
            mapid = create_update_mapping(
                conn,
                form_data["reference"],
                form_data["description"],
                target_schema,
                form_data["tableType"],
                form_data["tableName"],
                form_data["freqCode"],
                form_data["sourceSystem"],
                "N",  # lgvrfyflg
                datetime.datetime.now(),
                "N",  # stflg
                form_data["bulkProcessRows"],
                p_trgconid=target_connection_id,
                p_user=user_id,
                p_chkpntstrtgy=checkpoint_strategy,
                p_chkpntclnm=checkpoint_column,
                p_chkpntenbld=checkpoint_enabled,
            )

            # Optionally persist base/source SQL text for this mapping.
            base_sql = form_data.get("baseSql")
            if base_sql:
                try:
                    cur2 = conn.cursor()
                    db_type = _detect_db_type_from_connection(conn)
                    if db_type == "POSTGRESQL":
                        # Preferred: BASESQL (new name)
                        try:
                            cur2.execute(
                                "UPDATE DMS_MAPR SET BASESQL = %s WHERE MAPID = %s",
                                (base_sql, int(mapid)),
                            )
                        except Exception:
                            # Backward-compat: fallback to SRCSQL if BASESQL not present
                            cur2.execute(
                                "UPDATE DMS_MAPR SET SRCSQL = %s WHERE MAPID = %s",
                                (base_sql, int(mapid)),
                            )
                    else:
                        try:
                            cur2.execute(
                                "UPDATE DMS_MAPR SET BASESQL = :base_sql WHERE MAPID = :mapid",
                                {"base_sql": base_sql, "mapid": int(mapid)},
                            )
                        except Exception:
                            cur2.execute(
                                "UPDATE DMS_MAPR SET SRCSQL = :base_sql WHERE MAPID = :mapid",
                                {"base_sql": base_sql, "mapid": int(mapid)},
                            )
                    cur2.close()
                except Exception as e:
                    # Do not fail the save if the column is not present yet
                    warning(
                        f"save_to_db: unable to persist base SQL for mapping {form_data.get('reference')}: {e}"
                    )

            processed_rows: List[Dict[str, Any]] = []
            for idx, row in enumerate(rows):
                if not row.get("fieldName", "").strip():
                    continue

                # Respect modifiedRows: skip unmodified rows with existing mapdtlid
                if not (idx in modified_rows or not row.get("mapdtlid")):
                    continue

                logic_ver_flag = row.get("LogicVerFlag", "")

                mapdtlid = create_update_mapping_detail(
                    conn,
                    form_data["reference"],
                    row["fieldName"],
                    row["dataType"],
                    "Y" if row.get("primaryKey") else "N",
                    row.get("pkSeq") if row.get("pkSeq") and row.get("primaryKey") else None,
                    row.get("fieldDesc"),
                    row["logic"] if row.get("logic", "").strip() else None,
                    row.get("keyColumn"),
                    row.get("valColumn"),
                    row.get("mapCombineCode"),
                    row.get("execSequence"),
                    row.get("scdType"),
                    logic_ver_flag,
                    "" if logic_ver_flag == "" else datetime.datetime.now(),
                    form_data.get("username"),
                )

                processed_rows.append(
                    {"index": idx, "mapdtlid": mapdtlid, "fieldName": row["fieldName"]}
                )

            conn.commit()

            return {
                "success": True,
                "message": "Mapping saved successfully",
                "mapperId": str(mapid),
                "processedRows": processed_rows,
            }
        except Exception as e:
            if conn:
                conn.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"An error occurred while saving the mapping data: {str(e)}",
            )
        finally:
            if conn:
                conn.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"An error occurred while saving the mapping data: {str(e)}",
        )


class ExtractSqlColumnsRequest(BaseModel):
    sql_code: Optional[str] = None
    sql_content: Optional[str] = None
    connection_id: Optional[int] = None


class ExtractedColumn(BaseModel):
    column_name: str
    source_data_type: Optional[str] = None
    source_precision: Optional[int] = None
    source_scale: Optional[int] = None
    suggested_data_type: Optional[str] = None
    suggested_data_type_options: List[str] = []
    nullable: Optional[bool] = None
    is_primary_key: Optional[bool] = None


class ExtractSqlColumnsResponse(BaseModel):
    success: bool
    message: str
    columns: List[ExtractedColumn]
    source_table: Optional[str] = None
    source_schema: Optional[str] = None
    sql_content: str


@router.post("/extract-sql-columns", response_model=ExtractSqlColumnsResponse)
async def extract_sql_columns(payload: ExtractSqlColumnsRequest):
    """
    Given a SQL code or raw SQL, execute a lightweight wrapped query to discover
    column metadata and suggest target data types using the parameter table.

    This endpoint is designed for the mapper SQL prefill dialog.
    """
    data = payload.model_dump()
    sql_code = data.get("sql_code")
    sql_content = data.get("sql_content")
    connection_id = data.get("connection_id")

    if not sql_code and not sql_content:
        raise HTTPException(
            status_code=400,
            detail="Either sql_code or sql_content must be provided",
        )

    metadata_conn = None
    source_conn = None
    try:
        # Step 1: Resolve SQL content (use Manage SQL table if only code is provided)
        metadata_conn = create_metadata_connection()
        if not metadata_conn:
            raise HTTPException(
                status_code=500, detail="Failed to create metadata connection"
            )

        db_type_meta = _detect_db_type_from_connection(metadata_conn)

        if sql_code and not sql_content:
            cursor = metadata_conn.cursor()
            try:
                if db_type_meta == "POSTGRESQL":
                    query = (
                        "SELECT MAPRSQL, SQLCONID "
                        "FROM DMS_MAPRSQL WHERE MAPRSQLCD = %s AND CURFLG = 'Y'"
                    )
                    cursor.execute(query, (sql_code,))
                else:
                    query = (
                        "SELECT MAPRSQL, SQLCONID "
                        "FROM DMS_MAPRSQL WHERE MAPRSQLCD = :sql_code AND CURFLG = 'Y'"
                    )
                    cursor.execute(query, {"sql_code": sql_code})

                row = cursor.fetchone()
                if not row:
                    raise HTTPException(
                        status_code=404,
                        detail=f"No SQL logic found for code: {sql_code}",
                    )

                sql_value = row[0]
                # Handle CLOB vs plain string
                sql_content = (
                    sql_value.read() if hasattr(sql_value, "read") else str(sql_value)
                )
                # If connection_id not explicitly provided, try to use SQLCONID
                if connection_id is None and row[1] is not None:
                    try:
                        connection_id = int(row[1])
                    except (TypeError, ValueError):
                        connection_id = None
            finally:
                cursor.close()

        if not sql_content:
            raise HTTPException(
                status_code=400, detail="Resolved SQL content is empty"
            )

        # Step 2: Create source connection (target or metadata)
        if connection_id is not None:
            try:
                source_conn = create_target_connection(connection_id)
            except Exception as e:
                error(
                    f"extract_sql_columns: failed to create target connection "
                    f"(connection_id={connection_id}): {str(e)}"
                )
                raise HTTPException(
                    status_code=500,
                    detail=(
                        f"Failed to connect to selected database for SQL analysis "
                        f"(connection_id={connection_id})"
                    ),
                )
        else:
            source_conn = metadata_conn

        if not source_conn:
            raise HTTPException(
                status_code=500, detail="Failed to create source connection"
            )

        db_type_source = _detect_db_type_from_connection(source_conn)

        # Step 3: Build wrapped SQL and execute to obtain column metadata
        wrapped_sql = _build_wrapped_sql(sql_content, db_type_source)
        info(
            f"extract_sql_columns: executing wrapped SQL for metadata only "
            f"(db_type={db_type_source})"
        )

        src_cursor = source_conn.cursor()
        try:
            src_cursor.execute(wrapped_sql)
            description = src_cursor.description or []
        finally:
            src_cursor.close()

        if not description:
            return ExtractSqlColumnsResponse(
                success=False,
                message="No columns found in SQL result",
                columns=[],
                source_table=None,
                source_schema=None,
                sql_content=sql_content,
            )

        # Step 4: Fetch all available target data types from parameter table
        datatype_rows = get_parameter_mapping_datatype(metadata_conn)
        # dataTypeOptions are dictionaries with keys like PRCD, PRDESC, PRVAL
        # We expose PRCD as the canonical code
        all_type_codes = [str(row.get("PRCD")) for row in datatype_rows if row.get("PRCD")]

        # Build parsed option buckets using both PRCD (code) and PRVAL (database type)
        # to drive "closest fit" suggestions.
        import re

        def _parse_int(value) -> Optional[int]:
            try:
                return int(value)
            except (TypeError, ValueError):
                return None

        # Buckets
        string_buckets: list[tuple[int, str]] = []  # (length, PRCD)
        numeric_buckets: list[tuple[int, int, str]] = []  # (precision, scale, PRCD)
        generic_string_codes: list[str] = []
        integer_like_codes: list[str] = []
        bigint_like_codes: list[str] = []
        decimal_like_codes: list[str] = []  # Generic DECIMAL/NUMERIC/NUMBER without explicit p,s
        date_like_codes: list[str] = []
        timestamp_like_codes: list[str] = []
        time_like_codes: list[str] = []

        for row in datatype_rows:
            code_raw = row.get("PRCD")
            val_raw = row.get("PRVAL")
            if not code_raw:
                continue
            code = str(code_raw)
            val = str(val_raw) if val_raw is not None else ""
            cu = code.upper()
            vu = val.upper()

            # Identify date/time
            if "TIMESTAMP" in cu or "TIMESTAMP" in vu:
                timestamp_like_codes.append(code)
            if "DATE" in cu or "DATE" in vu:
                date_like_codes.append(code)
            if " TIME" in cu or " TIME" in vu:  # avoid matching 'TIMESTAMP' here
                time_like_codes.append(code)

            # Identify generic families
            if any(tok in cu for tok in ["INTEGER", " INT"]):  # ' INT' avoids capturing 'BIGINT' here
                if "BIGINT" not in cu:
                    integer_like_codes.append(code)
            if "BIGINT" in cu:
                bigint_like_codes.append(code)
            if any(tok in cu for tok in ["DECIMAL", "NUMERIC", "NUMBER"]):
                decimal_like_codes.append(code)
            if any(tok in cu for tok in ["CHAR", "VARCHAR", "STRING", "TEXT", "CLOB", "NCHAR", "NVARCHAR"]):
                generic_string_codes.append(code)

            # Parse PRVAL patterns
            # String length patterns, e.g. VARCHAR2(30), VARCHAR(50), CHAR(10)
            m_len = re.search(r"(VARCHAR2|VARCHAR|CHAR|NCHAR|NVARCHAR|TEXT|CLOB)\s*\(\s*(\d+)\s*\)", vu)
            if m_len:
                length = _parse_int(m_len.group(2))
                if length:
                    string_buckets.append((length, code))

            # Numeric patterns, e.g. DECIMAL(10,2), NUMBER(22), NUMERIC(18,0)
            m_num = re.search(r"(NUMBER|NUMERIC|DECIMAL)\s*\(\s*(\d+)\s*(?:,\s*(\d+)\s*)?\)", vu)
            if m_num:
                prec = _parse_int(m_num.group(2))
                sca = _parse_int(m_num.group(3)) if m_num.group(3) is not None else 0
                if prec is not None:
                    numeric_buckets.append((prec, sca or 0, code))
                continue  # done with PRVAL parsing

            # Parse PRCD patterns as a fallback, e.g. String20, Number22, Decimal10_2, "money 10"
            # StringN
            m_code_len = re.search(r"(STRING|CHAR|VARCHAR|NCHAR|NVARCHAR)\s*[_ ]?\s*(\d+)$", cu)
            if m_code_len:
                length = _parse_int(m_code_len.group(2))
                if length:
                    string_buckets.append((length, code))

            # Decimal10_2 / Number22 / Money 10
            m_code_num = re.search(r"(DECIMAL|NUMERIC|NUMBER|MONEY)\s*(_|\s)?\s*(\d+)(?:[_\s,]+(\d+))?$", cu)
            if m_code_num:
                prec = _parse_int(m_code_num.group(3))
                sca = _parse_int(m_code_num.group(4)) if m_code_num.group(4) is not None else 0
                if prec is not None:
                    numeric_buckets.append((prec, sca or 0, code))

        def suggest_type_for_source(
            source_type: str,
            source_precision: Optional[int],
            source_scale: Optional[int],
        ) -> Optional[str]:
            """
            Heuristic mapping from source type + precision/scale to one of the
            available target type codes (PRCD) from DMS_PARAMS (PRTYP='Datatype').

            This is designed to work well with common PRCD conventions like:
            - String5, String20, String255
            - Integer, BigInt, Number, Decimal10_2
            - Date, Timestamp
            """
            if not source_type or not all_type_codes:
                return None

            st = str(source_type).upper()
            sp = source_precision
            ss = source_scale

            # Build quick lookup structures once per call
            codes_upper = [(c, str(c).upper()) for c in all_type_codes]

            # 1) Date / time
            if any(tok in st for tok in ["DATE", "TIMESTAMP", "TIME"]):
                # Prefer TIMESTAMP when source indicates it
                if "TIMESTAMP" in st:
                    for c, cu in codes_upper:
                        if "TIMESTAMP" in cu or cu in ("TS",):
                            return c
                # Otherwise prefer DATE
                for c, cu in codes_upper:
                    if cu == "DATE" or "DATE" in cu:
                        return c
                # Fallback: any time-like option
                for c, cu in codes_upper:
                    if any(k in cu for k in ["TIME", "TS"]):
                        return c

            # 2) Character / string (use length/precision to choose nearest bucket like String20)
            if any(tok in st for tok in ["CHAR", "VARCHAR", "TEXT", "CLOB", "STRING", "NCHAR", "NVARCHAR"]):
                if string_buckets and isinstance(sp, int) and sp > 0:
                    # Choose the smallest StringN that can fit (>= length). If none, choose max.
                    string_buckets_sorted = sorted(string_buckets, key=lambda x: x[0])
                    for n, code in string_buckets_sorted:
                        if n >= sp:
                            return code
                    return string_buckets_sorted[-1][1]

                # No size buckets available - fallback to a generic string-ish code
                for code in generic_string_codes:
                    return code
                for c, cu in codes_upper:
                    if any(k in cu for k in ["STRING", "CHAR", "VARCHAR", "TEXT", "CLOB", "NCHAR", "NVARCHAR"]):
                        return c

            # 3) Numeric
            if any(tok in st for tok in ["NUMBER", "NUMERIC", "DECIMAL", "INT", "BIGINT", "SMALLINT", "FLOAT", "DOUBLE", "MONEY"]):
                # If scale exists and is > 0, prefer a DECIMAL/NUMERIC style code
                if isinstance(ss, int) and ss > 0:
                    # Find a numeric bucket with both precision and scale that can fit
                    candidates = [
                        (p, s, code) for (p, s, code) in numeric_buckets
                        if p is not None and p >= (sp or 0) and s is not None and s >= ss
                    ]
                    if candidates:
                        p, s, code = sorted(candidates, key=lambda x: (x[0], x[1]))[0]
                        return code
                    # Fallback to any decimal-like code
                    for code in decimal_like_codes:
                        return code

                # Integer-like if scale is 0/None
                if isinstance(sp, int) and sp > 0:
                    # Try integer/bigint preferences based on typical thresholds
                    if sp <= 9 and integer_like_codes:
                        return integer_like_codes[0]
                    if sp <= 18 and bigint_like_codes:
                        return bigint_like_codes[0]

                    # If there are precision-based numeric buckets like Number22/Decimal22, choose nearest >=
                    candidates = [
                        (p, s, code) for (p, s, code) in numeric_buckets
                        if p is not None and p >= sp and (s is None or s == 0)
                    ]
                    if candidates:
                        p, s, code = sorted(candidates, key=lambda x: (x[0], x[1]))[0]
                        return code

                # Last resort: any numeric-ish code
                for code in decimal_like_codes or integer_like_codes or bigint_like_codes:
                    return code
                for c, cu in codes_upper:
                    if any(k in cu for k in ["NUM", "DEC", "INT", "NUMBER", "BIGINT"]):
                        return c

            # 4) Fallback: no suggestion
            return None

        # Step 5: Build column list (including precision/scale where available)
        columns: List[ExtractedColumn] = []
        for col in description:
            # Cursor description structure is driver-specific; we only rely on name and type object
            col_name = str(col[0]) if col and col[0] is not None else ""

            source_type_str: Optional[str] = None
            source_precision: Optional[int] = None
            source_scale: Optional[int] = None

            # Common pattern: description is a sequence with type object at index 1
            type_obj = col[1] if len(col) > 1 else None
            if type_obj is not None:
                source_type_str = str(type_obj)
                # Try Oracle-style type object attributes
                if hasattr(type_obj, "precision"):
                    try:
                        source_precision = int(type_obj.precision) if type_obj.precision is not None else None
                    except (TypeError, ValueError):
                        source_precision = None
                if hasattr(type_obj, "scale"):
                    try:
                        source_scale = int(type_obj.scale) if type_obj.scale is not None else None
                    except (TypeError, ValueError):
                        source_scale = None
                # For character types, some drivers expose length via .size
                if source_precision is None and hasattr(type_obj, "size"):
                    try:
                        size_val = getattr(type_obj, "size", None)
                        if isinstance(size_val, int) and size_val > 0:
                            source_precision = size_val
                    except Exception:
                        pass

            # Psycopg2-style: precision/scale at fixed positions (4, 5)
            if source_precision is None and len(col) > 4:
                try:
                    prec_val = col[4]
                    source_precision = int(prec_val) if prec_val is not None else None
                except (TypeError, ValueError):
                    source_precision = None
            if source_scale is None and len(col) > 5:
                try:
                    scale_val = col[5]
                    source_scale = int(scale_val) if scale_val is not None else None
                except (TypeError, ValueError):
                    source_scale = None

            # Fallback for character types where precision is not populated:
            # use display/internal size from the cursor description if available.
            if (
                source_precision is None
                and source_type_str
                and any(tok in source_type_str.upper() for tok in ["CHAR", "VARCHAR", "TEXT", "CLOB", "NCHAR", "NVARCHAR"])
            ):
                # Many DB-API descriptions are: (name, type_code, display_size, internal_size, precision, scale, null_ok)
                # Try display_size (index 2) or internal_size (index 3).
                try:
                    if len(col) > 2 and isinstance(col[2], int) and col[2] > 0:
                        source_precision = col[2]
                    elif len(col) > 3 and isinstance(col[3], int) and col[3] > 0:
                        source_precision = col[3]
                except Exception:
                    pass

            suggested_type = suggest_type_for_source(
                source_type_str or "",
                source_precision,
                source_scale,
            )

            columns.append(
                ExtractedColumn(
                    column_name=col_name,
                    source_data_type=source_type_str,
                    source_precision=source_precision,
                    source_scale=source_scale,
                    suggested_data_type=suggested_type,
                    suggested_data_type_options=all_type_codes,
                    nullable=None,
                    is_primary_key=None,
                )
            )

        return ExtractSqlColumnsResponse(
            success=True,
            message=f"Successfully extracted {len(columns)} columns from SQL",
            columns=columns,
            source_table=None,
            source_schema=None,
            sql_content=sql_content,
        )
    except HTTPException:
        raise
    except Exception as e:
        error(f"extract_sql_columns: unexpected error: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"An error occurred while extracting SQL columns: {str(e)}",
        )
    finally:
        # Only close metadata_conn if it's not the same as source_conn
        try:
            if metadata_conn and metadata_conn is not source_conn:
                metadata_conn.close()
        except Exception:
            pass
        # If source_conn is a separate target connection, close it
        try:
            if source_conn and source_conn is not metadata_conn:
                source_conn.close()
        except Exception:
            pass


class CheckSqlDuplicateRequest(BaseModel):
    sql_content: str
    connection_id: Optional[int] = None
    similarity_threshold: float = 0.7


class SimilarQuery(BaseModel):
    sql_code: str
    similarity_score: float
    sql_content: str


class CheckSqlDuplicateResponse(BaseModel):
    has_exact_match: bool
    exact_match_code: Optional[str] = None
    similar_queries: List[SimilarQuery] = []


@router.post("/check-sql-duplicate", response_model=CheckSqlDuplicateResponse)
async def check_sql_duplicate(payload: CheckSqlDuplicateRequest):
    """
    Check if a given SQL already exists (exactly or similarly) in DMS_MAPRSQL.

    Uses simplified normalization + string similarity (difflib) and returns
    any matches above the provided similarity threshold (default: 0.7).
    """
    data = payload.model_dump()
    sql_content = (data.get("sql_content") or "").strip()
    similarity_threshold = float(data.get("similarity_threshold") or 0.7)

    if not sql_content:
        raise HTTPException(
            status_code=400, detail="sql_content must not be empty"
        )

    normalized_new = _normalize_sql(sql_content)
    if not normalized_new:
        raise HTTPException(
            status_code=400,
            detail="Normalized SQL content is empty; please provide a valid SQL",
        )

    conn = None
    try:
        conn = create_metadata_connection()
        if not conn:
            raise HTTPException(
                status_code=500, detail="Failed to create metadata connection"
            )

        db_type = _detect_db_type_from_connection(conn)
        cursor = conn.cursor()
        try:
            if db_type == "POSTGRESQL":
                query = (
                    "SELECT MAPRSQLCD, MAPRSQL "
                    "FROM DMS_MAPRSQL WHERE CURFLG = 'Y'"
                )
                cursor.execute(query)
            else:
                query = (
                    "SELECT MAPRSQLCD, MAPRSQL "
                    "FROM DMS_MAPRSQL WHERE CURFLG = 'Y'"
                )
                cursor.execute(query)

            rows = cursor.fetchall() or []
        finally:
            cursor.close()

        has_exact_match = False
        exact_match_code: Optional[str] = None
        similar: List[SimilarQuery] = []

        for row in rows:
            code = str(row[0])
            sql_value = row[1]
            existing_sql = (
                sql_value.read() if hasattr(sql_value, "read") else str(sql_value)
            )
            normalized_existing = _normalize_sql(existing_sql)
            if not normalized_existing:
                continue

            if normalized_existing == normalized_new:
                has_exact_match = True
                exact_match_code = code
                # Even if exact match is found, continue to collect other similar queries

            score = _calculate_sql_similarity(normalized_new, normalized_existing)
            if score >= similarity_threshold:
                similar.append(
                    SimilarQuery(
                        sql_code=code,
                        similarity_score=round(score, 4),
                        sql_content=existing_sql,
                    )
                )

        # Sort similar queries by descending similarity
        similar_sorted = sorted(
            similar, key=lambda x: x.similarity_score, reverse=True
        )

        return CheckSqlDuplicateResponse(
            has_exact_match=has_exact_match,
            exact_match_code=exact_match_code,
            similar_queries=similar_sorted[:5],
        )
    except HTTPException:
        raise
    except Exception as e:
        error(f"check_sql_duplicate: unexpected error: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"An error occurred while checking SQL duplicates: {str(e)}",
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


class ValidateLogicRequest(BaseModel):
    p_logic: str
    p_keyclnm: str
    p_valclnm: str
    # Optional: validate against a specific target connection instead of metadata DB
    connection_id: Optional[int] = None
    # Optional: mapping reference to look up target connection if connection_id not provided
    mapref: Optional[str] = None


@router.post("/validate-logic")
async def validate_logic(payload: ValidateLogicRequest):
    """
    Validate a single piece of mapping logic.
    Mirrors Flask endpoint: POST /mapper/validate-logic
    """
    data = payload.model_dump()
    p_logic = data.get("p_logic")
    p_keyclnm = data.get("p_keyclnm")
    p_valclnm = data.get("p_valclnm")
    connection_id = data.get("connection_id")
    mapref = data.get("mapref")  # Get mapping reference to look up target connection

    if not all([p_logic, p_keyclnm, p_valclnm]):
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing required parameters. Please provide p_logic, "
                "p_keyclnm, and p_valclnm."
            ),
        )

    metadata_connection = None
    target_connection = None
    try:
        # Always create metadata connection for querying DMS_MAPRSQL table
        try:
            metadata_connection = create_metadata_connection()
            if not metadata_connection:
                raise HTTPException(
                    status_code=500,
                    detail="Failed to create metadata connection"
                )
        except Exception as e:
            error(f"Error creating metadata connection: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to create metadata connection: {str(e)}"
            )
        
        # Determine target connection ID:
        # 1. First, use connection_id if explicitly provided
        # 2. Otherwise, look up target connection from mapping reference if mapref is provided
        trgconid = None
        info(f"validate_logic: Received connection_id={connection_id}, mapref={mapref}")
        if connection_id is not None and connection_id != 0:
            trgconid = connection_id
            info(f"Using provided connection_id={connection_id} for validation")
        elif mapref:
            # Look up target connection from mapping reference
            try:
                from backend.modules.helper_functions import get_mapping_ref
                info(f"Looking up target connection for mapref={mapref}")
                mapping_data = get_mapping_ref(metadata_connection, mapref)
                if mapping_data:
                    trgconid = mapping_data.get("TRGCONID") or mapping_data.get("trgconid")
                    info(f"Mapping data retrieved: TRGCONID={trgconid}, keys={list(mapping_data.keys())}")
                    if trgconid:
                        info(f"Found target connection ID {trgconid} from mapping reference {mapref}")
                    else:
                        warning(f"No target connection (TRGCONID) configured for mapping reference {mapref}")
                else:
                    warning(f"Mapping reference {mapref} not found in database")
            except Exception as e:
                error(f"Error looking up target connection for mapref {mapref}: {str(e)}", exc_info=True)
                # Continue without target connection
        
        # Create target connection if we have a connection ID
        if trgconid:
            try:
                info(f"Creating target connection for validation (connection_id={trgconid})")
                target_connection = create_target_connection(trgconid)
                if not target_connection:
                    error(f"create_target_connection returned None for connection_id={trgconid}")
                    # Don't fail - just log and continue with metadata connection
                else:
                    # Verify target connection can access the database
                    try:
                        test_cursor = target_connection.cursor()
                        # Detect database type from connection module
                        import builtins
                        module_name = builtins.type(target_connection).__module__
                        if "psycopg" in module_name or "pg8000" in module_name:
                            test_cursor.execute("SELECT current_database(), current_schema()")
                        else:
                            test_cursor.execute("SELECT sys_context('userenv', 'db_name'), sys_context('userenv', 'current_schema') FROM dual")
                        test_result = test_cursor.fetchone()
                        test_cursor.close()
                        info(f"Successfully created target connection (ID: {trgconid}) for SQL validation. Database info: {test_result}")
                    except Exception as test_e:
                        warning(f"Target connection created but test query failed: {str(test_e)}")
                    info(f"Target connection (ID: {trgconid}) ready for SQL validation")
            except Exception as e:
                error(f"Failed to create target connection (ID: {trgconid}): {str(e)}", exc_info=True)
                # Don't fail the entire request - just log and continue with metadata connection
                # The validation will use metadata connection instead
                target_connection = None
        else:
            warning("No target connection_id available, will use metadata connection for SQL validation (tables may not exist)")
        
        # Pass both connections: metadata for metadata queries, target for SQL validation
        # Log connection details before validation
        info(f"About to call validate_logic2: target_connection is {'NOT None' if target_connection else 'None'}")
        if target_connection:
            try:
                import builtins
                target_module = builtins.type(target_connection).__module__
                metadata_module = builtins.type(metadata_connection).__module__
                info(f"Connection types - Metadata: {metadata_module}, Target: {target_module}")
                # Check if they're the same object
                if target_connection is metadata_connection:
                    error("CRITICAL: target_connection and metadata_connection are the SAME object!")
                else:
                    info("Connection objects are different (good)")
            except Exception as conn_check_e:
                warning(f"Could not verify connection objects: {str(conn_check_e)}")
        
        try:
            is_valid, err = validate_logic2(metadata_connection, p_logic, p_keyclnm, p_valclnm, target_connection)
            return {
                "status": "success",
                "is_valid": is_valid,
                "message": "Logic is valid" if is_valid == "Y" else err,
            }
        except Exception as validation_error:
            error(f"Error in validate_logic2: {str(validation_error)}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Error during SQL validation: {str(validation_error)}"
            )
    except HTTPException:
        raise
    except Exception as e:
        error(f"Unexpected error in validate_logic endpoint: {str(e)}")
        error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500, detail=f"Error validating logic: {str(e)}"
        )
    finally:
        if metadata_connection:
            try:
                metadata_connection.close()
            except Exception:
                pass
        if target_connection:
            try:
                target_connection.close()
            except Exception:
                pass


class ValidateBatchRequest(BaseModel):
    mapref: str
    rows: List[Dict[str, Any]] = []


@router.post("/validate-batch")
async def validate_batch_logic(payload: ValidateBatchRequest):
    """
    Validate all mapping details for a given mapref in bulk.
    Mirrors Flask endpoint: POST /mapper/validate-batch
    """
    data = payload.model_dump()
    p_mapref = data.get("mapref")
    rows = data.get("rows", [])

    metadata_connection = None
    target_connection = None
    try:
        metadata_connection = create_metadata_connection()

        # Determine target connection from mapping
        mapping_data = get_mapping_ref(metadata_connection, p_mapref)
        if mapping_data:
            trgconid = mapping_data.get("TRGCONID") or mapping_data.get("trgconid")
        else:
            trgconid = None

        if trgconid:
            try:
                target_connection = create_target_connection(trgconid)
            except Exception:
                # Fallback to metadata connection if target fails
                target_connection = metadata_connection
        else:
            target_connection = metadata_connection

        results: List[Dict[str, Any]] = []

        # Bulk validation using helper function
        bulk_result, bulk_error = validate_all_mapping_details(
            metadata_connection, p_mapref, target_connection
        )

        if bulk_error is None:
            for row in rows:
                if row.get("logic"):
                    results.append(
                        {
                            "rowId": row.get("mapdtlid"),
                            "fieldName": row.get("fieldName"),
                            "isValid": True,
                            "error": None,
                            "detailedError": "Logic is Verified",
                        }
                    )
            return {
                "status": "success",
                "bulkValidation": {"success": True, "error": None},
                "rowResults": results,
            }

        # Collect all mapdtlids that have logic
        map_detail_ids = [
            row.get("mapdtlid")
            for row in rows
            if row.get("mapdtlid") and row.get("logic")
        ]

        error_messages = get_error_messages_list(metadata_connection, map_detail_ids)

        for row in rows:
            if not row.get("logic"):
                continue

            err_msg = None
            if row.get("mapdtlid") and row.get("mapdtlid") in error_messages:
                err_msg = error_messages[row.get("mapdtlid")]

            results.append(
                {
                    "rowId": row.get("mapdtlid"),
                    "fieldName": row.get("fieldName"),
                    "isValid": err_msg == "Logic is Verified",
                    "error": None if err_msg == "Logic is Verified" else err_msg,
                    "detailedError": err_msg,
                }
            )

        return {
            "status": "success",
            "bulkValidation": {"success": bulk_result, "error": bulk_error},
            "rowResults": results,
        }
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error in validate_batch: {str(e)}"
        )
    finally:
        if metadata_connection:
            metadata_connection.close()
        if target_connection and target_connection is not metadata_connection:
            try:
                target_connection.close()
            except Exception:
                pass


class ActivateDeactivateRequest(BaseModel):
    mapref: str
    statusFlag: str


@router.post("/activate-deactivate")
async def activate_deactivate_mapping(payload: ActivateDeactivateRequest):
    """
    Activate or deactivate a mapping.
    Mirrors Flask endpoint: POST /mapper/activate-deactivate
    """
    data = payload.model_dump()
    p_mapref = data.get("mapref")
    p_stflg = data.get("statusFlag")

    if not p_mapref or not p_stflg:
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing required parameters. Please provide mapref and statusFlag."
            ),
        )

    if p_stflg not in ["A", "N"]:
        raise HTTPException(
            status_code=400,
            detail='Invalid status flag. Must be either "A" (activate) or "N" (deactivate).',
        )

    conn = None
    try:
        conn = create_metadata_connection()
        success, message = call_activate_deactivate_mapping(conn, p_mapref, p_stflg)
        return {"success": success, "message": message}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"An error occurred while processing the request: {str(e)}",
        )
    finally:
        if conn:
            conn.close()


# ----- Reference management & delete endpoints -----


@router.get("/get-all-mapper-reference")
async def get_all_mapper_reference() -> List[List[Any]]:
    """
    Get all mapper reference details.
    Mirrors Flask endpoint: GET /mapper/get-all-mapper-reference
    """
    conn = None
    try:
        conn = create_metadata_connection()
        query = """
        SELECT MAPREF, MAPDESC,TRGSCHM,TRGTBTYP,FRQCD,SRCSYSTM,LGVRFYFLG,STFLG,CRTDBY,UPTDBY
        FROM DMS_MAPR
        WHERE CURFLG = 'Y'
        """
        cursor = conn.cursor()
        cursor.execute(query)
        result = cursor.fetchall()
        cursor.close()
        # Convert tuples to lists for JSON serialization
        return [list(row) for row in result]
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error in get_all_mapper_reference: {str(e)}",
        )
    finally:
        if conn:
            conn.close()


class DeleteMapperReferenceRequest(BaseModel):
    mapref: str


@router.post("/delete-mapper-reference")
async def delete_mapper_reference(payload: DeleteMapperReferenceRequest):
    """
    Delete a mapper reference.
    Mirrors Flask endpoint: POST /mapper/delete-mapper-reference
    """
    conn = None
    try:
        p_mapref = payload.mapref
        conn = create_metadata_connection()

        try:
            success, message = call_delete_mapping(conn, p_mapref)

            if success:
                return {"success": True, "message": message}
            else:
                raise HTTPException(
                    status_code=400,
                    detail={"success": False, "message": message},
                )
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error in delete_mapper_reference: {str(e)}",
        )


class DeleteMappingDetailRequest(BaseModel):
    mapref: str
    trgclnm: str


@router.post("/delete-mapping-detail")
async def delete_mapping_detail(payload: DeleteMappingDetailRequest):
    """
    Delete a mapping detail row.
    Mirrors Flask endpoint: POST /mapper/delete-mapping-detail
    """
    conn = None
    try:
        p_mapref = payload.mapref
        p_trgclnm = payload.trgclnm

        if not p_mapref or not p_trgclnm:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Missing required parameters. Please provide mapref and trgclnm."
                ),
            )

        conn = create_metadata_connection()
        try:
            success, message = call_delete_mapping_details(conn, p_mapref, p_trgclnm)

            return {"success": success, "message": message}
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"An error occurred while deleting the mapping detail: {str(e)}",
        )


# ----- Template/Excel endpoints -----


@router.get("/download-template")
async def download_template(format: str = Query("xlsx", regex="^(xlsx|csv)$")):
    """
    Download an empty mapper template (Excel or CSV).
    Mirrors Flask endpoint: GET /mapper/download-template?format=xlsx|csv
    """
    try:
        export_format = format.lower()
        if export_format not in ["xlsx", "csv"]:
            export_format = "xlsx"

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping Template"

        # Define styles
        header_fill = PatternFill(
            start_color="00B050", end_color="00B050", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write Form Fields section
        ws["A1"] = "Form Fields"
        ws.merge_cells("A1:H1")
        ws["A1"].fill = header_fill
        ws["A1"].font = header_font
        ws["A1"].alignment = header_alignment

        # Write Form Fields headers
        for col, field in enumerate(FORM_FIELDS, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Add empty row for form data
        for col in range(1, len(FORM_FIELDS) + 1):
            cell = ws.cell(row=3, column=col)
            cell.border = border

        # Add space between sections
        ws.append([])

        # Write Table Fields section
        current_row = 5
        ws.cell(row=current_row, column=1, value="Table Fields")
        ws.merge_cells(f"A{current_row}:K{current_row}")
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = header_alignment

        # Write Table Fields headers
        current_row += 1
        for col, field in enumerate(TABLE_FIELDS, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Add empty rows for table data
        for row in range(current_row + 1, current_row + 11):  # 10 empty rows
            for col in range(1, len(TABLE_FIELDS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

        # Adjust column widths
        for col_idx in range(1, max(len(FORM_FIELDS), len(TABLE_FIELDS)) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            # Check form headers (row 2)
            if col_idx <= len(FORM_FIELDS):
                header_value = ws.cell(row=2, column=col_idx).value
                if header_value:
                    max_length = max(max_length, len(str(header_value)))

            # Check table headers (row 6)
            if col_idx <= len(TABLE_FIELDS):
                header_value = ws.cell(row=current_row, column=col_idx).value
                if header_value:
                    max_length = max(max_length, len(str(header_value)))

            # Set the column width
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()

        if export_format == "csv":
            # Create DataFrame from workbook data to save as CSV
            form_headers = [
                ws.cell(row=2, column=i).value
                for i in range(1, len(FORM_FIELDS) + 1)
            ]
            form_data = [
                ws.cell(row=3, column=i).value or ""
                for i in range(1, len(FORM_FIELDS) + 1)
            ]

            table_headers = [
                ws.cell(row=current_row, column=i).value
                for i in range(1, len(TABLE_FIELDS) + 1)
            ]
            table_data = []
            for row in range(current_row + 1, current_row + 11):
                row_data = []
                for col in range(1, len(TABLE_FIELDS) + 1):
                    row_data.append(ws.cell(row=row, column=col).value or "")
                table_data.append(row_data)

            # First save form data
            form_df = pd.DataFrame([form_data], columns=form_headers)

            # Then save table data
            table_df = pd.DataFrame(table_data, columns=table_headers)

            # Combine them with a separator row
            combined_csv = form_df.to_csv(index=False) + "\n\n" + table_df.to_csv(
                index=False
            )

            # Convert to bytes and prepare response
            excel_buffer = io.BytesIO(combined_csv.encode())
            mimetype = "text/csv"
            download_name = "mapper_template.csv"
        else:  # xlsx format
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            mimetype = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            download_name = "mapper_template.xlsx"

        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type=mimetype,
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error in download_template: {str(e)}"
        )


class DownloadCurrentRequest(BaseModel):
    formData: Dict[str, Any]
    rows: List[Dict[str, Any]]
    format: Optional[str] = "xlsx"


@router.post("/download-current")
async def download_current(payload: DownloadCurrentRequest):
    """
    Download current mapping data as Excel or CSV.
    Mirrors Flask endpoint: POST /mapper/download-current
    """
    try:
        data = payload.model_dump()
        form_data = data.get("formData", {})
        rows_data = data.get("rows", [])

        # Check if format parameter is provided, default to xlsx
        export_format = data.get("format", "xlsx").lower()
        if export_format not in ["xlsx", "csv"]:
            export_format = "xlsx"

        # Filter out rows without field names
        rows_data = [row for row in rows_data if row.get("fieldName")]

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping Template"

        # Define styles
        header_fill = PatternFill(
            start_color="00B050", end_color="00B050", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write Form Fields section
        ws["A1"] = "Form Fields"
        ws.merge_cells("A1:H1")
        ws["A1"].fill = header_fill
        ws["A1"].font = header_font
        ws["A1"].alignment = header_alignment

        # Write Form Fields headers
        for col, field in enumerate(FORM_FIELDS, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Write Form Fields values
        for col, field in enumerate(FORM_FIELDS, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = form_data.get(field, "")
            cell.border = border

        # Add space between sections
        ws.append([])

        # Write Table Fields section
        current_row = 5
        ws.cell(row=current_row, column=1, value="Table Fields")
        ws.merge_cells(f"A{current_row}:K{current_row}")
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = header_alignment

        # Write Table Fields headers
        current_row += 1
        for col, field in enumerate(TABLE_FIELDS, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = field
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Add rows with data
        for row_idx, row_data in enumerate(rows_data, current_row + 1):
            # Map the field names to match TABLE_FIELDS
            field_mapping = {
                "primaryKey": "primaryKey",
                "pkSeq": "pkSeq",
                "fieldName": "fieldName",
                "dataType": "dataType",
                "fieldDesc": "fieldDesc",
                "scdType": "scdType",
                "keyColumn": "keyColumn",
                "valColumn": "valColumn",
                "logic": "logic",
                "mapCombineCode": "mapCombineCode",
                "execSequence": "execSequence",
            }

            for col_idx, field in enumerate(TABLE_FIELDS, 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Get the mapped field name
                mapped_field = field_mapping.get(field, field)

                # Handle boolean fields like primaryKey separately
                if field == "primaryKey":
                    cell.value = "Yes" if row_data.get(mapped_field, False) else "No"
                else:
                    cell.value = row_data.get(mapped_field, "")

                cell.border = border

        # Adjust column widths
        for col_idx in range(1, max(len(FORM_FIELDS), len(TABLE_FIELDS)) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            # Check all rows for maximum value length
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    cell_value_str = str(cell_value)
                    # Limit preview length for very long values
                    max_length = max(max_length, min(len(cell_value_str), 50))

            # Set the column width
            adjusted_width = max(max_length + 2, 12)  # Minimum width of 12
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()

        if export_format == "csv":
            # Create DataFrame from workbook data to save as CSV
            form_headers = [
                ws.cell(row=2, column=i).value
                for i in range(1, len(FORM_FIELDS) + 1)
            ]
            form_data_values = [
                ws.cell(row=3, column=i).value or ""
                for i in range(1, len(FORM_FIELDS) + 1)
            ]

            table_headers = [
                ws.cell(row=current_row, column=i).value
                for i in range(1, len(TABLE_FIELDS) + 1)
            ]
            table_data = []
            for row in range(current_row + 1, current_row + 1 + len(rows_data)):
                row_data = []
                for col in range(1, len(TABLE_FIELDS) + 1):
                    row_data.append(ws.cell(row=row, column=col).value or "")
                table_data.append(row_data)

            # First save form data
            form_df = pd.DataFrame([form_data_values], columns=form_headers)

            # Then save table data
            table_df = pd.DataFrame(table_data, columns=table_headers)

            # Combine them with a separator row
            combined_csv = form_df.to_csv(index=False) + "\n\n" + table_df.to_csv(
                index=False
            )

            # Convert to bytes and prepare response
            excel_buffer = io.BytesIO(combined_csv.encode())
            mimetype = "text/csv"
            download_name = f"{form_data.get('reference', 'mapper')}_data.csv"
        else:  # xlsx format
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            mimetype = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            download_name = f"{form_data.get('reference', 'mapper')}_data.xlsx"

        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type=mimetype,
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error in download_current: {str(e)}"
        )


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload and parse a mapper template Excel/CSV file.
    Mirrors Flask endpoint: POST /mapper/upload
    """
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file selected")

        # Read the file content
        file_content = await file.read()

        # Read the Excel file
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # Process form fields
        form_data = {}
        form_headers = [cell.value for cell in ws[2] if cell.value]
        form_values = [
            cell.value for cell in ws[3] if cell.column <= len(form_headers)
        ]

        for header, value in zip(form_headers, form_values):
            form_data[header] = str(value) if value is not None else ""

        # Process table fields
        table_start_row = 6  # Table headers start at row 6
        table_headers = [cell.value for cell in ws[table_start_row] if cell.value]
        rows = []

        for row in ws.iter_rows(min_row=table_start_row + 1, max_col=len(table_headers)):
            row_data = {}
            has_data = False

            for header, cell in zip(table_headers, row):
                value = cell.value

                # Handle boolean fields
                if header == "primaryKey":
                    if isinstance(value, bool):
                        value = value
                    elif isinstance(value, (int, float)):
                        value = bool(value)
                    else:
                        value = (
                            str(value).lower().strip() in ["true", "1", "yes", "y"]
                            if value
                            else False
                        )
                elif value is None:
                    value = ""
                else:
                    value = str(value).strip()
                    if value:
                        has_data = True

                row_data[header] = value

            if has_data:
                rows.append(row_data)

        # Map the data to the required format
        mapped_rows = []
        for row in rows:
            mapped_row = {
                "mapdtlid": "",  # This will be empty for new rows
                "fieldName": row.get("fieldName", ""),
                "dataType": row.get("dataType", ""),
                "primaryKey": row.get("primaryKey", False),
                "pkSeq": row.get("pkSeq", ""),
                "nulls": False,  # Default value
                "logic": row.get("logic", ""),
                "validator": "N",  # Default value
                "keyColumn": row.get("keyColumn", ""),
                "valColumn": row.get("valColumn", ""),
                "mapCombineCode": row.get("mapCombineCode", ""),
                "LogicVerFlag": "N",  # Default value
                "scdType": row.get("scdType", ""),
                "fieldDesc": row.get("fieldDesc", ""),
                "execSequence": row.get("execSequence", ""),
            }
            mapped_rows.append(mapped_row)

        # Prepare response data
        response_data = {
            "formData": {
                "reference": str(form_data.get("reference", "")).strip(),
                "description": str(form_data.get("description", "")).strip(),
                "mapperId": "",  # This might need to be generated or extracted
                "targetSchema": str(form_data.get("targetSchema", "")).strip(),
                "tableName": str(form_data.get("tableName", "")).strip(),
                "tableType": str(form_data.get("tableType", "")).strip(),
                "freqCode": str(form_data.get("freqCode", "")).strip(),
                "sourceSystem": str(form_data.get("sourceSystem", "")).strip(),
                "bulkProcessRows": form_data.get("bulkProcessRows", ""),
            },
            "rows": mapped_rows,
        }

        return response_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error in upload_file: {str(e)}")




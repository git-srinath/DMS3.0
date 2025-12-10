"""
DEPRECATED: This Flask application is no longer in use.
All functionality has been migrated to FastAPI (see backend/fastapi_app.py).
This file is kept for reference only and should not be run.
The application now runs exclusively on FastAPI via uvicorn.
"""

from flask import Flask, request, jsonify, send_file, g
from flask_cors import CORS
import pandas as pd
import uuid
import os
import datetime
import json
import io
import oracledb
from sqlalchemy import create_engine, text
import random
import re
import traceback
from dotenv import load_dotenv
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.dbconnect import create_oracle_connection, sqlite_engine

app = Flask(__name__)
# CORS(app, resources={
#     r"/*": {
#         "origins": ["http://localhost:3000"],
#         "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
#         "allow_headers": ["Content-Type", "Authorization"],
#         "supports_credentials": True,
#         "expose_headers": ["Authorization"]
#     }
# })

# allow all origins with  support credentials
# CORS(app, resources={
#     r"/*": {
#         "origins": ["*"],
#         "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
#         "supports_credentials": True,
#     }
# })

CORS(app, supports_credentials=True)
# Load environment variables
load_dotenv()

# Error handler for all exceptions
@app.errorhandler(Exception)
def handle_exception(e):
    # Import logger inside the function to avoid circular imports
    from modules.logger import error
    # Log the exception without traceback
    error(f"Unhandled exception: {str(e)}")
    return jsonify({"error": "An unexpected error occurred"}), 500

# No blueprints registered - all functionality migrated to FastAPI

# Create directories if they don't exist
os.makedirs('data/drafts', exist_ok=True)
os.makedirs('data/templates', exist_ok=True)

if __name__ == '__main__':
    # Import logger inside the function to avoid circular imports
    from modules.logger import info
    info("Server starting...")
    app.run(debug=True, port=5000, host='0.0.0.0')   
    info("Server stopped.")
 